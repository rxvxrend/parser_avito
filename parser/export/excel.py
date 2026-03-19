from pathlib import Path
from threading import Lock

from openpyxl import Workbook, load_workbook

from dto import ParsedProduct
from parser.export.base import ResultStorage


class ExcelStorage(ResultStorage):
    name = "excel"
    headers = ["Ссылка", "Название", "Цена", "Описание"]

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = Lock()

        if not self.file_path.exists():
            self._create_file()

    def _create_file(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append(self.headers)
        workbook.save(self.file_path)

    @staticmethod
    def excel_safe(value):
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value

    def save(self, items: list[ParsedProduct]) -> None:
        if not items:
            return

        with self._lock:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active

            for item in items:
                sheet.append([
                    self.excel_safe(item.url),
                    self.excel_safe(item.title),
                    self.excel_safe(item.price),
                    self.excel_safe(item.description),
                ])

            workbook.save(self.file_path)
